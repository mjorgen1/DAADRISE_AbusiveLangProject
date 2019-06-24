# extract German Twitter data here from txt file and save to new excel sheet

# need to install openpyxl via pip (if you want to use via command line) or you can install via pycharm
from openpyxl import Workbook

def main():

    # create a new excel sheet
    wb = Workbook()

    # select the sheet we want to use and name it
    ws =  wb.active
    ws.title = "German Tweet Data"

    # adding headers to the sheet
    ws['A1'] = "Comment_ID"
    ws['B1'] = "Comment"
    ws['C1'] = "Label"

    row_num=2
    comm_ID_num = 1

    # opens the text file
    f=open('germeval2018.training.txt','r')

    # iterates through the text file line by line
    for line in f.readlines():
        # print(line)
        comment_ID = 'A' + str(row_num)
        ws[comment_ID] = comm_ID_num

        # TODO: find first instance of non " " char string in line
        # TODO: concatenate the chars following that instance until a "   " is found
        comment = ""
        comment_place = 'B' + str(row_num)
        ws[comment_place] = comment

        # TODO: iterate through the next word of chars until the next " " is found
        # TODO: save the following chars to a string which is then added to C_ in the table
        label = ""
        label_place = 'C' + str(row_num)
        ws[label_place] = label

        #update the row_num for the next line and the commID num
        row_num += 1
        comm_ID_num += 1

        # this if is here for the purposes of testing
        if row_num == 2:
            break

    # save the excel file in the current dataExtraction folder
    wb.save(filename = 'excelTest.xlsx')

if __name__ == '__main__':
    main()
